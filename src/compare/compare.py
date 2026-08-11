import pandas as pd
import os
import glob
import configparser
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
import pymysql
from pymysql import Error

from src.compare.read_company import read_company_with_name
from src.compare.label_input import main as input_main

BASE_DIR = Path(__file__).resolve().parent.parent.parent


# ================= 日期转换辅助函数 =================
def convert_period_to_date(period_str):
    """
    将时间段字符串转换为该时间段的第一天的日期
    例如："25_8_25_9" -> "2025-08-01"
    """
    if not period_str or pd.isna(period_str):
        return None

    try:
        # 分割时间段字符串
        parts = period_str.split('_')
        if len(parts) != 4:
            print(f"警告: 时间段格式不正确: {period_str}")
            return None

        # 提取开始年份和月份
        start_year_short = parts[0]  # 25
        start_month = parts[1]  # 8

        # 将短年份转换为完整年份
        if len(start_year_short) == 2:
            # 假设20xx年
            start_year = int(f"20{start_year_short}")
        else:
            start_year = int(start_year_short)

        # 转换月份
        start_month = int(start_month)

        # 创建日期对象，取该月的第一天
        date_obj = datetime(start_year, start_month, 1)

        # 返回字符串格式的日期
        return date_obj.strftime('%Y-%m-%d')

    except Exception as e:
        print(f"警告: 无法转换时间段 {period_str} 为日期: {e}")
        return None


# ================= 配置读取函数 =================
def load_label_config(config_path):
    """从配置文件加载标签分析设置"""
    config = configparser.ConfigParser()

    # 设置默认值
    config['FOLDERS'] = {
        '结构点_score_col': 'point_risk',
        '结构点_keys': 'point_name,direction',
        '路段_score_col': 'line_risk',
        '路段_keys': 'line,direction',
        '路网_score_col': 'net_risk',
        '路网_keys': 'net_comprehensive'
    }

    config['PERIODS'] = {
        'periods': '25_6_25_7,25_8_25_9,25_10_25_11,25_12_26_1'
    }

    config['THRESHOLDS'] = {
        'up_threshold_min': '70',
        'up_threshold_max': '80',
        'middle_risk_threshold': '80',
        'high_risk_threshold': '90'
    }

    config['OUTPUT'] = {
        'output_file': '风险评价对比结果.xlsx'
    }

    # 读取配置文件
    if os.path.exists(config_path):
        config.read(config_path, encoding='utf-8')
        print(f"✓ 已加载标签配置文件: {config_path}")
    else:
        print(f"⚠ 标签配置文件 {config_path} 不存在，使用默认配置")

    return config


def get_config_value(config, section, key, default=None, data_type=str):
    """安全获取配置值"""
    try:
        value = config.get(section, key, fallback=default)
        if data_type == int:
            return int(value)
        elif data_type == float:
            return float(value)
        elif data_type == bool:
            return config.getboolean(section, key, fallback=default)
        elif data_type == list:
            return [item.strip() for item in value.split(',')] if value else []
        else:
            return value
    except (ValueError, configparser.NoSectionError, configparser.NoOptionError) as e:
        print(f"警告: 读取配置 [{section}]{key} 失败: {e}, 使用默认值: {default}")
        return default


def create_risk_contrast_table(db_config):
    """创建风险对比表"""
    try:
        # 连接数据库
        connection = pymysql.connect(
            host=db_config['host'],
            port=db_config['port'],
            user=db_config['user'],
            password=db_config['password'],
            database=db_config['database'],
            charset='utf8mb4'
        )

        with connection.cursor() as cursor:
            # 创建表的SQL语句，添加asso_company字段
            create_table_sql = f"""
            CREATE TABLE IF NOT EXISTS `{db_config['table_name']}` (
                `id` BIGINT PRIMARY KEY COMMENT '期次编码(YYMM+序号)',
                `belong_date` DATE NOT NULL COMMENT '归属日期',
                `unit_type` VARCHAR(255) COMMENT '单元分类（点线网）',
                `structure_name` VARCHAR(255) NOT NULL COMMENT '结构名称',
                `asso_company` VARCHAR(255) COMMENT '所属公司',
                `direction` VARCHAR(255) NOT NULL COMMENT '上下行',
                `risk` DOUBLE COMMENT '当期风险值',
                `risk_level` VARCHAR(255) COMMENT '风险等级',
                `pre_risk` DOUBLE COMMENT '上期风险值',
                `change_type` VARCHAR(255) COMMENT '风险演化类型',  
                `created_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY `unique_structure` (`structure_name`, `direction`, `belong_date`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci COMMENT='风险对比分析表';
            """

            # 执行创建表
            cursor.execute(create_table_sql)

            print(f"✓ 已检查/创建数据库表: {db_config['table_name']}")

        connection.commit()
        connection.close()

    except Error as e:
        print(f"❌ 数据库连接或表创建失败: {e}")
        return False

    return True


def insert_data_to_db(df, contrast_period, db_config):
    """将数据插入到数据库"""
    if df.empty:
        print(f"⚠ 对比周期 {contrast_period} 没有数据可插入数据库")
        return False

    # 转换对比周期为归属日期
    belong_date = convert_period_to_date(contrast_period)
    if not belong_date:
        print(f"❌ 无法转换对比周期 {contrast_period} 为有效日期")
        return False

    print(f"  转换对比周期: {contrast_period} -> 归属日期: {belong_date}")

    try:
        # 连接数据库
        connection = pymysql.connect(
            host=db_config['host'],
            port=db_config['port'],
            user=db_config['user'],
            password=db_config['password'],
            database=db_config['database'],
            charset='utf8mb4'
        )

        with connection.cursor() as cursor:
            # 先删除相同belong_date的旧数据，实现替换而非追加
            delete_sql = f"DELETE FROM `{db_config['table_name']}` WHERE belong_date = %s"
            cursor.execute(delete_sql, (belong_date,))
            deleted_count = cursor.rowcount
            if deleted_count > 0:
                print(f"  已删除 {deleted_count} 条相同归属日期的旧记录")

            # 生成期次前缀 YYMM
            dt = datetime.strptime(belong_date, '%Y-%m-%d')
            prefix = dt.strftime('%y%m')

            # 插入新数据
            insert_sql = f"""
            INSERT INTO `{db_config['table_name']}` 
            (id, belong_date, unit_type, structure_name, direction, risk, risk_level, pre_risk, change_type, asso_company)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """

            data_to_insert = []
            for idx, (_, row) in enumerate(df.iterrows(), start=1):
                unit_type = row['评估单元类型'] if '评估单元类型' in df.columns else row.get('unit_type', '')
                structure_name = row['结构名称'] if '结构名称' in df.columns else row.get('structure_name', '')
                direction = row['方向'] if '方向' in df.columns else row.get('direction', '')

                asso_company = read_company_with_name(unit_type, structure_name)

                record_id = int(f"{prefix}{idx:03d}")
                data_to_insert.append((
                    record_id,
                    belong_date,
                    unit_type,
                    structure_name,
                    direction,
                    float(row['当期风险值'] if '当期风险值' in df.columns else row.get('risk', 0)),
                    row['风险等级'] if '风险等级' in df.columns else row.get('risk_level', ''),
                    float(row['往期风险值'] if '往期风险值' in df.columns else row.get('pre_risk', 0)),
                    row['风险变化类型'] if '风险变化类型' in df.columns else row.get('change_type', ''),
                    asso_company
                ))

            cursor.executemany(insert_sql, data_to_insert)
            print(f"✓ 已插入 {len(data_to_insert)} 条记录到数据库")

        connection.commit()
        connection.close()
        return True

    except Error as e:
        print(f"❌ 数据库插入失败: {e}")
        return False


# ================= 核心逻辑 =================
def get_risk_label(p_prev, p_curr, thresholds):
    """判定风险变化标签逻辑"""
    UP_THRESHOLD_MIN = thresholds['up_threshold_min']
    UP_THRESHOLD_MAX = thresholds['up_threshold_max']
    MIDDLE_RISK_THRESHOLD = thresholds['middle_risk_threshold']
    HIGH_RISK_THRESHOLD = thresholds['high_risk_threshold']

    # 1. 风险上行：相邻两期风险值均在up_threshold区间内且为递增
    if (UP_THRESHOLD_MIN < p_prev <= UP_THRESHOLD_MAX) and \
            (UP_THRESHOLD_MIN < p_curr <= UP_THRESHOLD_MAX) and \
            (p_curr > p_prev):
        return "风险上行"
    # 2. 风险下行：相邻两期风险值均在up_threshold区间内且为递减
    elif (UP_THRESHOLD_MIN < p_prev <= UP_THRESHOLD_MAX) and \
            (UP_THRESHOLD_MIN < p_curr <= UP_THRESHOLD_MAX) and \
            (p_curr < p_prev):
        return "风险下行"
    # 3. 风险上行：往期在middle以下，当期进入一般风险区间(middle~high)
    elif p_prev < MIDDLE_RISK_THRESHOLD and MIDDLE_RISK_THRESHOLD <= p_curr < HIGH_RISK_THRESHOLD:
        return "风险上行"
    # 4. 当期高风险：往期在high以下，当期直接进入较高风险区间(>=high)
    elif p_prev < HIGH_RISK_THRESHOLD and p_curr >= HIGH_RISK_THRESHOLD:
        return "当期高风险"
    # 5. 持续高风险：相邻两期均在high以上
    elif p_prev >= HIGH_RISK_THRESHOLD and p_curr >= HIGH_RISK_THRESHOLD:
        return "持续高风险"
    # 6. 脱离高风险：往期在high以上但当期降至high以下
    elif p_prev >= HIGH_RISK_THRESHOLD and p_curr < HIGH_RISK_THRESHOLD:
        return "脱离高风险"
    else:
        return "无需重点关注"


def get_risk_level(score, thresholds):
    """根据分数获取风险等级"""
    if score < thresholds['middle_risk_threshold']:
        return "低风险"
    elif thresholds['middle_risk_threshold'] <= score < thresholds['high_risk_threshold']:
        return "一般风险"
    else:
        return "较高风险"


def find_file(folder, period_tag):
    """查找包含特定周期标签的Excel文件"""
    # 尝试多种可能的文件名格式
    patterns = [
        f"*{period_tag}*",  # 新格式: 25_6_25_7
        f"*{period_tag.replace('_', ',')}*",  # 旧格式: 25,6,25,7
        f"*{period_tag.replace('_', '、')}*",  # 可能的格式: 25、6、25、7
    ]

    for pattern in patterns:
        files = glob.glob(os.path.join(folder, pattern))
        if files:
            return files[0]

    # 如果没有找到，尝试另一种查找方式
    files = glob.glob(os.path.join(folder, "*.xlsx"))
    files.extend(glob.glob(os.path.join(folder, "*.xls")))

    # 尝试在文件名中查找周期标识
    for file in files:
        if period_tag in file:
            return file

    return None


def clean_column_names(df):
    """清理列名：去除空格和特殊字符"""
    df.columns = [str(col).strip() for col in df.columns]
    return df


def get_sheet_name_for_excel(period_str):
    """为Excel工作表生成名称: risk_contrast_25_12_26_1"""
    return f"risk_contrast_{period_str}"


def get_sheet_name_for_display(period_str):
    """为显示生成友好的工作表名称"""
    parts = period_str.split('_')
    if len(parts) == 4:
        year1 = f"20{parts[0]}"
        month1 = parts[1]
        year2 = f"20{parts[2]}"
        month2 = parts[3]

        # 如果月份有前导0，去掉
        month1 = str(int(month1))
        month2 = str(int(month2))

        return f"当期({year1}年{month1}月-{year2}年{month2}月)_vs_往期"
    else:
        return period_str


def process_data():
    # 加载对比分析配置
    compare_config_path = str(BASE_DIR / 'config' / 'compare.ini')

    label_config = load_label_config(compare_config_path)

    # 从统一 output_db.ini 获取数据库配置
    from src.db_config import get_output_db_config
    output_db = get_output_db_config()

    # 解析标签配置
    FOLDERS = {}
    for unit_type in ['结构点', '路段', '路网']:
        score_col = get_config_value(label_config, 'FOLDERS', f'{unit_type}_score_col')
        keys_str = get_config_value(label_config, 'FOLDERS', f'{unit_type}_keys', '')
        keys = keys_str.split(',') if keys_str else []

        FOLDERS[unit_type] = {
            "score_col": score_col,
            "keys": keys
        }

    # 周期配置
    periods_str = get_config_value(label_config, 'PERIODS', 'periods', '25_6_25_7,25_8_25_9,25_10_25_11,25_12_26_1')
    period_pairs = [p.strip() for p in periods_str.split(',')]
    PERIODS = period_pairs

    # 阈值配置
    THRESHOLDS = {
        'up_threshold_min': get_config_value(label_config, 'THRESHOLDS', 'up_threshold_min', 70, int),
        'up_threshold_max': get_config_value(label_config, 'THRESHOLDS', 'up_threshold_max', 80, int),
        'middle_risk_threshold': get_config_value(label_config, 'THRESHOLDS', 'middle_risk_threshold', 80, int),
        'high_risk_threshold': get_config_value(label_config, 'THRESHOLDS', 'high_risk_threshold', 90, int)
    }

    # 输出文件配置
    OUTPUT_FILE = 'data/output/' + get_config_value(label_config, 'OUTPUT', 'output_file', '风险评价对比结果.xlsx')

    # 数据库配置 (从统一 output_db.ini 获取)
    db_config_dict = {
        'host': output_db.get('host', 'localhost'),
        'port': output_db.get('port', 3306),
        'user': output_db.get('user', 'root'),
        'password': output_db.get('password', ''),
        'database': output_db.get('database', 'risk_analysis'),
        'table_name': output_db.get('risk_contrast_table', 'risk_contrast')
    }

    # 打印配置信息
    print("\n=== 配置信息 ===")
    print(f"评估单元类型: {list(FOLDERS.keys())}")
    print(f"对比周期: {PERIODS}")
    print(f"阈值配置: {THRESHOLDS}")
    print(f"输出文件: {OUTPUT_FILE}")
    print(f"数据库表: {db_config_dict['table_name']}")
    print("=" * 20 + "\n")

    # 创建数据库表
    if create_risk_contrast_table(db_config_dict):
        print("✓ 数据库表准备就绪")
    else:
        print("⚠ 数据库表创建失败，将跳过数据库入库操作")

    # 使用 openpyxl 引擎
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # 依次处理对比Sheet
        for i in range(len(PERIODS) - 1):
            prev_tag = PERIODS[i]
            curr_tag = PERIODS[i + 1]

            # 生成Excel工作表名称
            excel_sheet_name = get_sheet_name_for_excel(curr_tag)
            # 确保工作表名称不超过31个字符（Excel限制）
            if len(excel_sheet_name) > 31:
                excel_sheet_name = excel_sheet_name[:31]

            # 生成显示用名称
            display_name = get_sheet_name_for_display(curr_tag)

            all_data = []  # 存储所有需要关注的数据

            for unit_type, config_dict in FOLDERS.items():
                # 查找数据文件夹 (从数据库导出的文件存放位置)
                compare_folder = f'data/temp/compare/{unit_type}'
                prev_file = find_file(compare_folder, prev_tag)
                curr_file = find_file(compare_folder, curr_tag)

                if not prev_file or not curr_file:
                    print(f"提示：{unit_type} 缺少 {prev_tag} 或 {curr_tag} 的文件，已跳过。")
                    continue

                # 读取数据
                df_prev = pd.read_excel(prev_file)
                df_curr = pd.read_excel(curr_file)

                # 清理列名
                df_prev = clean_column_names(df_prev)
                df_curr = clean_column_names(df_curr)

                # 获取配置中的列名
                target_score_col = config_dict["score_col"]
                keys = config_dict["keys"]

                # 检查必需的列是否存在
                required_cols = keys + [target_score_col]
                missing_cols_prev = [col for col in required_cols if col not in df_prev.columns]
                missing_cols_curr = [col for col in required_cols if col not in df_curr.columns]

                if missing_cols_prev or missing_cols_curr:
                    print(f"警告：{unit_type} 文件中缺少必需列")
                    if missing_cols_prev:
                        print(f"  {os.path.basename(prev_file)} 缺少: {missing_cols_prev}")
                    if missing_cols_curr:
                        print(f"  {os.path.basename(curr_file)} 缺少: {missing_cols_curr}")
                    print(f"  已跳过 {unit_type} 的对比")
                    continue

                # 转换分数为数值
                df_prev[target_score_col] = pd.to_numeric(df_prev[target_score_col], errors='coerce').fillna(0)
                df_curr[target_score_col] = pd.to_numeric(df_curr[target_score_col], errors='coerce').fillna(0)

                # 提取往期关键得分用于对比
                df_prev_sub = df_prev[keys + [target_score_col]].rename(columns={target_score_col: "往期风险值"})

                # 以当期数据为主表进行合并
                merged = pd.merge(df_curr, df_prev_sub, on=keys, how="left").fillna(0)

                # 计算标签
                merged["风险变化类型"] = merged.apply(
                    lambda row: get_risk_label(row["往期风险值"], row[target_score_col], THRESHOLDS), axis=1
                )

                # 筛选出需要关注的样本
                filtered_df = merged[merged["风险变化类型"] != "无需重点关注"].copy()

                if not filtered_df.empty:
                    # 根据不同的单元类型映射字段
                    if unit_type == "结构点":
                        filtered_df["结构名称"] = filtered_df.get("point_name", f"{unit_type}_结构名称")
                        filtered_df["方向"] = filtered_df.get("direction", f"{unit_type}_方向")
                    elif unit_type == "路段":
                        filtered_df["结构名称"] = filtered_df.get("line", f"{unit_type}_结构名称")
                        filtered_df["方向"] = filtered_df.get("direction", f"{unit_type}_方向")
                    elif unit_type == "路网":
                        filtered_df["结构名称"] = filtered_df.get("net_comprehensive", f"{unit_type}_结构名称")
                        filtered_df["方向"] = "不区分"
                    else:
                        filtered_df["结构名称"] = f"{unit_type}_结构名称"
                        filtered_df["方向"] = f"{unit_type}_方向"

                    # 创建标准化的数据框
                    standard_df = pd.DataFrame({
                        "评估单元类型": unit_type,
                        "结构名称": filtered_df["结构名称"],
                        "方向": filtered_df["方向"],
                        "当期风险值": filtered_df[target_score_col],
                        "风险等级": filtered_df[target_score_col].apply(lambda x: get_risk_level(x, THRESHOLDS)),
                        "往期风险值": filtered_df["往期风险值"],
                        "风险变化类型": filtered_df["风险变化类型"]
                    })

                    all_data.append(standard_df)

            # 合并所有数据
            if all_data:
                combined_sheet = pd.concat(all_data, ignore_index=True)

                # 按当期风险值降序排序
                combined_sheet = combined_sheet.sort_values(by="当期风险值", ascending=False)

                # 写入Excel（使用规范的Excel工作表名称）
                combined_sheet.to_excel(writer, sheet_name=excel_sheet_name, index=False)

                # 获取工作簿和工作表对象进行格式设置
                workbook = writer.book
                worksheet = writer.sheets[excel_sheet_name]

                # 设置列宽
                column_widths = {
                    'A': 15,  # 评估单元类型
                    'B': 30,  # 结构名称
                    'C': 15,  # 方向
                    'D': 15,  # 当期风险值
                    'E': 15,  # 风险等级
                    'F': 15,  # 往期风险值
                    'G': 20  # 风险变化类型
                }

                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

                # 设置表头格式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                center_alignment = Alignment(horizontal="center", vertical="center")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 应用表头格式
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border

                # 设置数据区域格式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=7):
                    for cell in row:
                        cell.alignment = center_alignment
                        cell.border = border

                # 设置数值格式
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=4).number_format = '0.00'  # 当期风险值
                    worksheet.cell(row=row, column=6).number_format = '0.00'  # 往期风险值

                print(f"✓ 已生成 {display_name} 工作表，共 {len(combined_sheet)} 条记录")

                # 数据入库
                print(f"  正在将数据入库到 {db_config_dict['table_name']}...")
                if insert_data_to_db(combined_sheet, curr_tag, db_config_dict):
                    print(f"  ✓ 数据入库成功")
                else:
                    print(f"  ⚠ 数据入库失败")

            else:
                # 创建空表提示
                empty_df = pd.DataFrame({
                    "提示": ["本周期内未发现符合风险变化条件的样本"],
                    "评估单元类型": [""],
                    "结构名称": [""],
                    "方向": [""],
                    "当期风险值": [""],
                    "风险等级": [""],
                    "往期风险值": [""],
                    "风险变化类型": [""]
                })
                empty_df.to_excel(writer, sheet_name=excel_sheet_name, index=False)
                print(f"⚠  {display_name} 没有发现符合条件的记录")

    # 重新打开文件，添加汇总信息
    wb = load_workbook(OUTPUT_FILE)

    # 创建汇总工作表
    ws_summary = wb.create_sheet(title="汇总统计")

    # 写入汇总信息
    summary_data = []
    for i in range(len(PERIODS) - 1):
        prev_tag = PERIODS[i]
        curr_tag = PERIODS[i + 1]

        # 使用Excel工作表名称
        excel_sheet_name = get_sheet_name_for_excel(curr_tag)
        if len(excel_sheet_name) > 31:
            excel_sheet_name = excel_sheet_name[:31]

        if excel_sheet_name in wb.sheetnames:
            ws = wb[excel_sheet_name]
            # 统计不同风险类型的数量
            if ws.max_row > 1:  # 有数据的情况
                risk_counts = {}
                total_count = ws.max_row - 1  # 减去表头

                for row in range(2, ws.max_row + 1):
                    risk_type = ws.cell(row=row, column=7).value  # 风险变化类型列
                    if risk_type:
                        risk_counts[risk_type] = risk_counts.get(risk_type, 0) + 1

                # 在汇总中使用格式化后的周期名称
                prev_formatted = get_sheet_name_for_display(prev_tag)
                curr_formatted = get_sheet_name_for_display(curr_tag)
                summary_data.append({
                    "对比周期": f"{curr_formatted} vs {prev_formatted}",
                    "总记录数": total_count,
                    "风险上行": risk_counts.get("风险上行", 0),
                    "风险下行": risk_counts.get("风险下行", 0),
                    "当期高风险": risk_counts.get("当期高风险", 0),
                    "持续高风险": risk_counts.get("持续高风险", 0),
                    "脱离高风险": risk_counts.get("脱离高风险", 0)
                })

    if summary_data:
        summary_df = pd.DataFrame(summary_data)
        # 写入汇总数据
        for r_idx, row in enumerate(summary_df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)

        # 写入表头
        headers = ["对比周期", "总记录数", "风险上行", "风险下行", "当期高风险", "持续高风险", "脱离高风险"]
        for c_idx, header in enumerate(headers, 1):
            ws_summary.cell(row=1, column=c_idx, value=header)

        # 设置汇总表格式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用表头格式
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # 设置数据区域格式
        for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = border

        # 设置列宽（对比周期列加宽以容纳完整年份信息）
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 12
        ws_summary.column_dimensions['C'].width = 12
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 12
        ws_summary.column_dimensions['F'].width = 12
        ws_summary.column_dimensions['G'].width = 12

    wb.save(OUTPUT_FILE)
    wb.close()

    print(f"\n🎉 处理成功！生成文件：{OUTPUT_FILE}")
    print(f"📊 包含以下工作表：")
    for sheet in wb.sheetnames:
        print(f"   - {sheet}")

    return True


def run_compare():
    """两期对比分析入口函数（供外部 main.py 调用）"""
    # 步骤1: 从数据库导出各期评估数据
    input_main()
    # 步骤2: 执行两期对比分析
    FOLDERS = ['data/temp/compare/结构点', 'data/temp/compare/路段', 'data/temp/compare/路网']
    missing = [f for f in FOLDERS if not os.path.exists(f)]
    if missing:
        print(f"❌ 错误：文件夹 {missing} 不存在，请检查数据库导出是否成功。")
        return False
    process_data()
    return True


if __name__ == "__main__":
    run_compare()